if __name__ == '__main__':
    from crisscross.core_functions.megastructures import Megastructure
    from itertools import product
    from pathlib import Path
    import numpy as np
    import openpyxl
    from orthoseq_generator import helper_functions as hf
    from orthoseq_generator import sequence_computations as sc
    import matplotlib.pyplot as plt

    design_dir = Path('/Users/matt/Partners HealthCare Dropbox/Matthew Aquilina/Origami Crisscross Team Docs/Papers/hash_cad/supplementary_data/figure_5_designs_and_data')

    mega_2_1 = Megastructure(import_design_file=str(design_dir / 'valency_2-1_design.xlsx'))
    mega_2_6 = Megastructure(import_design_file=str(design_dir / 'valency_2-6_design.xlsx'))
    mega_4_7 = Megastructure(import_design_file=str(design_dir / 'valency_4-7_design.xlsx'))
    mega_7_3 = Megastructure(import_design_file=str(design_dir / 'valency_7-3_design.xlsx'))
    designs = {'2-1': mega_2_1, '2-6': mega_2_6, '4-7': mega_4_7, '7-3': mega_7_3}

    # load sequence pairs from the fig 5 assembly handle pairs (correctly ordered)
    wb = openpyxl.load_workbook('assembly_handle_pairs.xlsx')
    ws = wb[wb.sheetnames[0]]
    sequence_pairs = []
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip title + column headers
        sequence_pairs.append((str(row[1]), str(row[2])))
    print(f'Loaded {len(sequence_pairs)} sequence pairs from assembly_handle_pairs.xlsx')

    import pickle
    on_target_pkl = Path('on_target_energies.pkl')
    off_target_pkl = Path('off_target_energies.pkl')

    if on_target_pkl.exists() and off_target_pkl.exists():
        print('Loading cached energy data from pkl files...')
        with open(on_target_pkl, 'rb') as f:
            on_target_energies = pickle.load(f)
        with open(off_target_pkl, 'rb') as f:
            off_target_energies = pickle.load(f)
        print(f'Loaded on-target and off-target energies from cache.')
    else:
        hf.USE_LIBRARY = False
        on_target_energies = sc.compute_ontarget_energies(sequence_pairs)
        print('On-target energies for sequence pairs computed.')
        off_target_energies = sc.compute_offtarget_energies(sequence_pairs)
        print('Off-target energies for sequence pairs computed.')
        with open(on_target_pkl, 'wb') as f:
            pickle.dump(on_target_energies, f)
        with open(off_target_pkl, 'wb') as f:
            pickle.dump(off_target_energies, f)
        print('Saved energy data to pkl files.')

    # scale all energies by RT to make them dimensionless
    R = 0.001987  # kcal/(mol·K)
    T = 37 + 273.15  # K
    RT = R * T
    on_target_energies = tuple(e / RT if isinstance(e, np.ndarray) else np.array(e) / RT for e in on_target_energies)
    off_target_energies = {k: v / RT for k, v in off_target_energies.items()}
    print(f'Scaled all NUPACK energies by RT = {RT:.4f} kcal/mol to obtain dimensionless units.')

    def build_combination_matrices(megastructure):
        """
        Builds two matrices containing all pairwise handle vs antihandle comparisons
        across all 4 shift/flip orientations (matching the layout of precise_hamming_compute).

        For each (handle, antihandle) pair and slat_length positions, 4 orientations are generated:
          1. handle shifted left
          2. handle shifted right
          3. reversed handle shifted left
          4. reversed handle shifted right

        Returns combination_matrix_1 (handles), combination_matrix_2 (antihandles),
        handle_keys, and antihandle_keys for tracing back to slat IDs.
        """
        handle_dict, antihandle_dict = megastructure.get_bag_of_slat_handles(remove_blank_slats=True)
        slat_length = len(next(iter(handle_dict.values())))

        num_pairs = len(handle_dict) * len(antihandle_dict)
        single_combo = 4 * slat_length - 2  # 126 unique orientations (j=0 duplicates removed for right-shifts)
        total_combos = single_combo * num_pairs

        combination_matrix_1 = np.zeros((total_combos, slat_length), dtype=np.uint16)
        combination_matrix_2 = np.zeros((total_combos, slat_length), dtype=np.uint16)

        for pair_idx, ((hk, handle_slat), (ahk, antihandle_slat)) in enumerate(product(handle_dict.items(), antihandle_dict.items())):
            base = pair_idx * single_combo
            reversed_slat = handle_slat[::-1]
            for j in range(slat_length):
                # 1. normal handle, shifted left
                combination_matrix_1[base + j, :slat_length - j] = handle_slat[j:]
                # 2. normal handle, shifted right (skip j=0, duplicates orientation 1 j=0)
                if j != 0:
                    combination_matrix_1[base + slat_length + j - 1, j:] = handle_slat[:slat_length - j]
                # 3. reversed handle, shifted left
                combination_matrix_1[base + 2 * slat_length - 1 + j, :slat_length - j] = reversed_slat[j:]
                # 4. reversed handle, shifted right (skip j=0, duplicates orientation 3 j=0)
                if j != 0:
                    combination_matrix_1[base + 3 * slat_length - 2 + j, j:] = reversed_slat[:slat_length - j]
            # antihandle is tiled identically for all orientations of this pair
            combination_matrix_2[base:base + single_combo, :] = antihandle_slat

        return combination_matrix_1, combination_matrix_2, list(handle_dict.keys()), list(antihandle_dict.keys()), slat_length


    def validate_standard_loss(combo_1, combo_2, h_keys, ah_keys, slat_len, megastructure, fudge_dg=10):
        """Computes max bond count and effective bond count from combination matrices
        and compares against eqcorr2d reference values. Returns the effective loss."""
        num_pairs = len(h_keys) * len(ah_keys)

        # eqcorr2d reference values (compensated)
        valency_results = megastructure.get_parasitic_interactions()
        print(f'  eqcorr2d - max bond count: {valency_results["worst_match_score"]}, '
              f'effective bond count (Loss): {valency_results["mean_log_score"]:.6f}')

        # compute match counts from combination matrices
        matches_per_row = np.sum((combo_1 == combo_2) & (combo_1 != 0), axis=1)

        # build uncompensated histogram
        max_match = int(matches_per_row.max())
        hist = np.zeros(max_match + 1, dtype=np.int64)
        match_vals, match_counts = np.unique(matches_per_row, return_counts=True)
        hist[match_vals] = match_counts

        # compensate: subtract expected connection matches
        connection_counts, _ = megastructure.get_slat_match_counts()
        for match_val, count in connection_counts.items():
            if match_val > 1 and match_val <= max_match:
                hist[match_val] -= count

        # worst match score (highest non-zero bin after compensation)
        worst_match = int(np.max(np.nonzero(hist)[0]))

        # effective bond count (mean log score)
        hist_truncated = hist[:worst_match + 1]
        sum_score = np.sum(hist_truncated * np.exp(fudge_dg * np.arange(len(hist_truncated))))
        mean_log_score = np.log(sum_score / num_pairs) / fudge_dg

        print(f'  manual slat match compute - max bond count: {worst_match}, '
              f'effective bond count (Loss): {mean_log_score:.6f}')
        return mean_log_score

    def energy_weighted_loss(combo_1, combo_2, h_keys, ah_keys, pair_energies, fudge_dg=10):
        """Computes loss using actual per-handle on-target energies instead of a uniform fudge factor.
        Returns the effective interaction energy."""
        num_pairs = len(h_keys) * len(ah_keys)

        # energy lookup: handle ID 0 → 0 (no match), handle ID n → pair_energies[n-1]
        energy_lookup = np.zeros(len(pair_energies) + 1)
        energy_lookup[1:] = pair_energies  # ΔG values (negative for favorable binding)

        # for each row, sum the on-target energies of all matching handle positions
        match_mask = (combo_1 == combo_2) & (combo_1 != 0)
        energy_matrix = energy_lookup[combo_1] * match_mask
        energy_per_row = np.sum(energy_matrix, axis=1)

        # analogous to the standard exp(fudge_dg * k): use exp(-ΔG_total) since ΔG is negative
        sum_score = np.sum(np.exp(-energy_per_row))
        effective_energy = np.log(sum_score / num_pairs) / fudge_dg

        worst_energy = np.max(-energy_per_row) / fudge_dg

        print(f'  energy-weighted loss - worst interaction energy: {worst_energy:.4f}, '
              f'effective interaction energy: {effective_energy:.4f}')
        return effective_energy

    def full_energy_loss(combo_1, combo_2, h_keys, ah_keys, pair_energies, off_target_energy_dict, fudge_dg=10):
        """Computes loss using both on-target and off-target energies for all overlapping positions.
        Returns the effective interaction energy."""
        num_pairs = len(h_keys) * len(ah_keys)
        n_handles = len(pair_energies)

        # build a 2D energy lookup: full_energy[h_id, ah_id] → energy
        # ID 0 means empty position → 0 energy
        full_energy = np.zeros((n_handles + 1, n_handles + 1))
        # on-target (diagonal): handle i with antihandle i
        for i in range(n_handles):
            full_energy[i + 1, i + 1] = pair_energies[i]
        # off-target: handle i with antihandle j (i != j)
        ha_energies = off_target_energy_dict['antihandle_handle_energies']
        for i in range(n_handles):
            for j in range(n_handles):
                if i != j:
                    full_energy[i + 1, j + 1] = ha_energies[i, j]

        # vectorized lookup: energy at each position for each row
        energy_matrix = full_energy[combo_1, combo_2]
        energy_per_row = np.sum(energy_matrix, axis=1)

        # same scoring as energy_weighted_loss
        sum_score = np.sum(np.exp(-energy_per_row))
        effective_energy = np.log(sum_score / num_pairs) / fudge_dg

        worst_energy = np.max(-energy_per_row) / fudge_dg

        print(f'  full energy loss - worst interaction energy: {worst_energy:.4f}, '
              f'effective interaction energy: {effective_energy:.4f}')
        return effective_energy

    def nucleotide_match_loss(combo_1, combo_2, h_keys, ah_keys, seq_pairs):
        """Counts the total number of matching nucleotides across all overlapping handle positions
        for each orientation. Reports the worst case (max total matching nucleotides)."""
        n_seqs = len(seq_pairs)
        seq_len = len(seq_pairs[0][0])

        # precompute nucleotide match counts for all (handle_i, antihandle_j) pairs
        # nt_match_lookup[i, j] = number of complementary base pairs (A-T, G-C)
        complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
        nt_match_lookup = np.zeros((n_seqs + 1, n_seqs + 1), dtype=np.int32)  # ID 0 → 0 matches
        for i in range(n_seqs):
            for j in range(n_seqs):
                rev_anti = seq_pairs[j][1][::-1]
                best = 0
                # try all shifts: positive = handle shifted right, negative = handle shifted left
                for s in range(-(seq_len - 1), seq_len):
                    h_start = max(0, s)
                    a_start = max(0, -s)
                    overlap = seq_len - abs(s)
                    count = sum(
                        complement[c1.upper()] == c2.upper()
                        for c1, c2 in zip(seq_pairs[i][0][h_start:h_start + overlap], rev_anti[a_start:a_start + overlap])
                    )
                    best = max(best, count)
                nt_match_lookup[i + 1, j + 1] = best

        # vectorized lookup and sum per row
        nt_matches_per_row = np.sum(nt_match_lookup[combo_1, combo_2], axis=1)

        average_nt_matches = np.average(nt_matches_per_row)
        print(f'  nucleotide match loss - average nucleotide matches: {average_nt_matches}')
        return average_nt_matches

    def nucleotide_weighted_loss(combo_1, combo_2, h_keys, ah_keys, seq_pairs, fudge_dg=10, allow_shifts=True):
        """Like the standard loss but replaces fudge_dg * match_count with
        (fudge_dg / 7) * total_nucleotide_matches per row, so a perfect 7-nt handle match
        contributes the same as one unit of fudge_dg."""
        n_seqs = len(seq_pairs)
        seq_len = len(seq_pairs[0][0])
        num_pairs = len(h_keys) * len(ah_keys)
        shift_label = 'with shifts' if allow_shifts else 'no shifts'

        complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
        nt_match_lookup = np.zeros((n_seqs + 1, n_seqs + 1), dtype=np.float64)
        shifts = range(-(seq_len - 1), seq_len) if allow_shifts else [2]  # shift=2 skips tt linkers on each side
        for i in range(n_seqs):
            for j in range(n_seqs):
                rev_anti = seq_pairs[j][1][::-1]
                best = 0
                for s in shifts:
                    h_start = max(0, s)
                    a_start = max(0, -s)
                    overlap = seq_len - abs(s)
                    count = sum(
                        complement[c1.upper()] == c2.upper()
                        for c1, c2 in zip(seq_pairs[i][0][h_start:h_start + overlap], rev_anti[a_start:a_start + overlap])
                    )
                    best = max(best, count)
                nt_match_lookup[i + 1, j + 1] = best

        # sum nucleotide matches per row, then scale by fudge_dg / 7
        nt_per_row = np.sum(nt_match_lookup[combo_1, combo_2], axis=1)
        scaled_per_row = (fudge_dg / 7) * nt_per_row

        sum_score = np.sum(np.exp(scaled_per_row))
        effective_loss = np.log(sum_score / num_pairs) / fudge_dg

        worst_loss = np.max(scaled_per_row) / fudge_dg

        print(f'  nucleotide-weighted loss ({shift_label}) - worst: {worst_loss:.4f}, '
              f'effective: {effective_loss:.4f}')
        return effective_loss

    def nearest_neighbour_loss(combo_1, combo_2, h_keys, ah_keys, seq_pairs, fudge_dg=10, allow_shifts=True):
        """Like the nucleotide-weighted loss but uses the nearest-neighbour model:
        counts consecutive complementary dinucleotide pairs instead of individual nucleotides.
        A perfect 7-nt handle match has 6 NN pairs, so scaled by fudge_dg / 6."""
        n_seqs = len(seq_pairs)
        seq_len = len(seq_pairs[0][0])
        num_pairs = len(h_keys) * len(ah_keys)
        shift_label = 'with shifts' if allow_shifts else 'no shifts'

        complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
        nn_match_lookup = np.zeros((n_seqs + 1, n_seqs + 1), dtype=np.float64)
        shifts = range(-(seq_len - 1), seq_len) if allow_shifts else [2]  # shift=2 skips tt linkers on each side
        for i in range(n_seqs):
            for j in range(n_seqs):
                rev_anti = seq_pairs[j][1][::-1]
                best = 0
                for s in shifts:
                    h_start = max(0, s)
                    a_start = max(0, -s)
                    overlap = seq_len - abs(s)
                    if overlap < 2:
                        continue
                    h_sub = seq_pairs[i][0][h_start:h_start + overlap]
                    a_sub = rev_anti[a_start:a_start + overlap]
                    # count consecutive positions where both are complementary
                    count = 0
                    for k in range(overlap - 1):
                        if complement[h_sub[k].upper()] == a_sub[k].upper() and complement[h_sub[k + 1].upper()] == a_sub[k + 1].upper():
                            count += 1
                    best = max(best, count)
                nn_match_lookup[i + 1, j + 1] = best

        # sum NN matches per row, then scale by fudge_dg / 6
        nn_per_row = np.sum(nn_match_lookup[combo_1, combo_2], axis=1)
        scaled_per_row = (fudge_dg / 6) * nn_per_row

        sum_score = np.sum(np.exp(scaled_per_row))
        effective_loss = np.log(sum_score / num_pairs) / fudge_dg

        worst_loss = np.max(scaled_per_row) / fudge_dg

        print(f'  nearest-neighbour loss ({shift_label}) - worst: {worst_loss:.4f}, '
              f'effective: {effective_loss:.4f}')
        return effective_loss

    pair_energies = on_target_energies[0]

    # collect effective losses for plotting
    design_names = []
    standard_losses = []
    energy_weighted_losses = []
    full_energy_losses = []
    full_nt_losses = []
    full_nt_noshift_losses = []
    nn_losses = []
    nn_noshift_losses = []
    plain_hamming_losses = []

    print('======================')
    for name, mega in designs.items():
        print('Analyzing design:', name)
        combo_1, combo_2, h_keys, ah_keys, slat_len = build_combination_matrices(mega)
        print('Validating standard loss computation:')
        std_loss = validate_standard_loss(combo_1, combo_2, h_keys, ah_keys, slat_len, mega)
        print('Energy-weighted loss (on-target only):')
        ew_loss = energy_weighted_loss(combo_1, combo_2, h_keys, ah_keys, pair_energies)
        print('Full energy loss (on-target + off-target):')
        fe_loss = full_energy_loss(combo_1, combo_2, h_keys, ah_keys, pair_energies, off_target_energies)
        print('Nucleotide match loss:')
        nt_loss = nucleotide_match_loss(combo_1, combo_2, h_keys, ah_keys, sequence_pairs)
        print('Nucleotide-weighted loss (with shifts):')
        ntw_loss = nucleotide_weighted_loss(combo_1, combo_2, h_keys, ah_keys, sequence_pairs, allow_shifts=True)
        print('Nucleotide-weighted loss (no shifts):')
        ntw_noshift_loss = nucleotide_weighted_loss(combo_1, combo_2, h_keys, ah_keys, sequence_pairs, allow_shifts=False)
        print('Nearest-neighbour loss (with shifts):')
        nn_loss = nearest_neighbour_loss(combo_1, combo_2, h_keys, ah_keys, sequence_pairs, allow_shifts=True)
        print('Nearest-neighbour loss (no shifts):')
        nn_noshift_loss = nearest_neighbour_loss(combo_1, combo_2, h_keys, ah_keys, sequence_pairs, allow_shifts=False)

        design_names.append(name)
        standard_losses.append(std_loss)
        energy_weighted_losses.append(ew_loss)
        full_energy_losses.append(fe_loss)
        full_nt_losses.append(ntw_loss)
        full_nt_noshift_losses.append(ntw_noshift_loss)
        nn_losses.append(nn_loss)
        nn_noshift_losses.append(nn_noshift_loss)
        plain_hamming_losses.append(nt_loss)
        print('======================')

    # normalize all losses to design 2-1 (first entry) as fold increase
    all_losses = {
        'Standard Loss': standard_losses,
        'Nupack Energy Loss': energy_weighted_losses,
        'Loss with Nupack Off-Target Binding': full_energy_losses,
        'Direct Nucleotide Complementarity (with shifts)': plain_hamming_losses,
        'Nucleotide Complementarity Loss (with shifts)': full_nt_losses,
        'Nucleotide Complementarity Loss (no shifts)': full_nt_noshift_losses,
        'Nearest Neighbour Complementarity Loss (with shifts)': nn_losses,
        'Nearest Neighbour Complementarity Loss (no shifts)': nn_noshift_losses,
    }
    normalized_losses = {}
    for label, values in all_losses.items():
        baseline = values[0]  # design 2-1
        normalized_losses[label] = [v / baseline for v in values]

    # bar chart
    plt.rcParams.update({
        'font.family': 'Helvetica',
        'font.size': 12,
        'axes.labelsize': 16,
        'axes.titlesize': 16,
        'xtick.labelsize': 13,
        'ytick.labelsize': 13,
        'legend.fontsize': 11,
    })

    dark2 = plt.colormaps['Dark2']
    n_bars = len(all_losses)
    x = np.arange(len(design_names))
    group_width = 0.85
    gap = 0.01  # absolute gap between bars
    bar_width = (group_width - gap * (n_bars - 1)) / n_bars

    display_names = [n.replace('-', '.') for n in design_names]

    fig, ax = plt.subplots(figsize=(10, 5))
    step = bar_width + gap
    offsets = np.arange(n_bars) * step - (n_bars - 1) * step / 2
    for i, (label, values) in enumerate(normalized_losses.items()):
        ax.bar(x + offsets[i], values, bar_width,
               label=label, color=dark2(i / n_bars), edgecolor='black', linewidth=0.5)

    ax.axhline(y=1.0, color='grey', linestyle='--', linewidth=0.8)
    ax.set_xlabel('Square Design Variant (Loss value provided)')
    ax.set_ylabel('Fold increase over 2.1 design')
    ax.set_xticks(x)
    ax.set_xticklabels(display_names)
    ax.legend()

    fig.tight_layout()
    # fig.savefig('loss_comparison.png', dpi=300)
    plt.show()
