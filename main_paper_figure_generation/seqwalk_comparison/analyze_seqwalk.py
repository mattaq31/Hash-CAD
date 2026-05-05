"""Generate SeqWalk pairs and plot on-target vs off-target energies."""

from pathlib import Path
import random
import sys

import numpy as np
import openpyxl

PACKAGE_ROOT = Path(__file__).resolve().parents[2] / "crisscross_kit"
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))

from orthoseq_generator import helper_functions as hf
from orthoseq_generator import sequence_computations as sc


def strongest_offtarget_interaction(sequence_pairs, off_energies):
    matrix_specs = [
        ("handle_handle", off_energies["handle_handle_energies"], "handle", "handle"),
        (
            "antihandle_antihandle",
            off_energies["antihandle_antihandle_energies"],
            "antihandle",
            "antihandle",
        ),
        (
            "handle_antihandle",
            off_energies["antihandle_handle_energies"],
            "handle",
            "antihandle",
        ),
    ]

    best = None
    for interaction_type, matrix, strand_type_a, strand_type_b in matrix_specs:
        valid_positions = np.argwhere(matrix != 0)
        for i, j in valid_positions:
            energy = matrix[i, j]
            if best is None or energy < best["energy"]:
                seq_a, rc_a = sequence_pairs[i]
                seq_b, rc_b = sequence_pairs[j]
                strand_a = seq_a if strand_type_a == "handle" else rc_a
                strand_b = seq_b if strand_type_b == "handle" else rc_b
                best = {
                    "interaction_type": interaction_type,
                    "energy": float(energy),
                    "pair_a_index": int(i),
                    "pair_b_index": int(j),
                    "pair_a": (seq_a, rc_a),
                    "pair_b": (seq_b, rc_b),
                    "strand_a_type": strand_type_a,
                    "strand_b_type": strand_type_b,
                    "strand_a": strand_a,
                    "strand_b": strand_b,
                }
    return best


def write_sequence_pairs_sheet(workbook, sheet_name, sequence_pairs):
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet["A1"] = sheet_name
    worksheet["A2"] = "ID"
    worksheet["B2"] = "Handle"
    worksheet["C2"] = "Antihandle"

    for idx, (seq, rc_seq) in enumerate(sequence_pairs, start=1):
        row = idx + 2
        worksheet.cell(row=row, column=1, value=idx)
        worksheet.cell(row=row, column=2, value=seq)
        worksheet.cell(row=row, column=3, value=rc_seq)


if __name__ == "__main__":
    SEQ_LENGTH = 7
    RANDOM_SEED = 41
    XLIM = (-13.5, 0.0)
    output_dir = Path(
        "/Users/floriankatzmeier/Dropbox/CrissCross/Papers/hash_cad/"
        "Figures/final_figures/revision/figure_SB"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "seqwalk_and_random_sequence_pairs.xlsx"
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    hf.ENERGY_TYPE = "totalu"
    hf.set_nupack_params(material="dna", celsius=37, sodium=0.05, magnesium=0.025)

    for k_value in range(3, 6):
        output_plot = output_dir / f"seqwalk_7mer_on_off_target_hist_k{k_value}.svg"
        random_output_plot = output_dir / f"random_7mer_on_off_target_hist_k{k_value}.svg"

        sequence_pairs_pool = sc.create_seqwalk_sequence_pairs_pool(
            length=SEQ_LENGTH,
            k=k_value,
            seed=RANDOM_SEED,
            alphabet="ACGT",
            avoid_reverse_complements=True,
            prevented_patterns=[],
            verbose=True,
        )

        print(
            f"Generated {len(sequence_pairs_pool)} SeqWalk sequence pairs "
            f"for {SEQ_LENGTH}-mers at SSM k={k_value}."
        )

        sequence_pairs = [pair for _, pair in sequence_pairs_pool]
        sequence_count = len(sequence_pairs)

        on_energies, _, _ = sc.compute_ontarget_energies(sequence_pairs)
        off_energies = sc.compute_offtarget_energies(sequence_pairs)

        stats = sc.plot_on_off_target_histograms(
            on_energies,
            off_energies,
            output_path=str(output_plot),
            show_plot=True,
            title=f"{sequence_count} SeqWalk Pairs (k={k_value})",
            xlim=XLIM,
        )
        strongest_offtarget = strongest_offtarget_interaction(sequence_pairs, off_energies)

        print(f"Plot saved to: {output_plot}")
        print("Worst off-target interaction:")
        print(
            f"{strongest_offtarget['interaction_type']} | "
            f"energy = {strongest_offtarget['energy']:.3f} kcal/mol"
        )
        print(
            f"Pair A index {strongest_offtarget['pair_a_index']}: "
            f"{strongest_offtarget['pair_a'][0]} / {strongest_offtarget['pair_a'][1]}"
        )
        print(
            f"Pair B index {strongest_offtarget['pair_b_index']}: "
            f"{strongest_offtarget['pair_b'][0]} / {strongest_offtarget['pair_b'][1]}"
        )
        print(
            f"Interacting strands: "
            f"{strongest_offtarget['strand_a_type']}={strongest_offtarget['strand_a']} "
            f"vs {strongest_offtarget['strand_b_type']}={strongest_offtarget['strand_b']}"
        )
        print(stats)
        write_sequence_pairs_sheet(workbook, f"k={k_value}", sequence_pairs)

        random.seed(RANDOM_SEED + k_value)
        random_pairs_pool = sc.create_sequence_pairs_pool(
            length=SEQ_LENGTH,
            fivep_ext="",
            threep_ext="",
            avoid_gggg=False,
        )
        random_sequence_pairs = sc.select_subset(
            random_pairs_pool,
            max_size=sequence_count,
        )

        random_on_energies, _, _ = sc.compute_ontarget_energies(random_sequence_pairs)
        random_off_energies = sc.compute_offtarget_energies(random_sequence_pairs)

        random_stats = sc.plot_on_off_target_histograms(
            random_on_energies,
            random_off_energies,
            output_path=str(random_output_plot),
            show_plot=True,
            title=f"{sequence_count} Random Sequence Pairs",
            xlim=XLIM,
        )
        random_strongest_offtarget = strongest_offtarget_interaction(
            random_sequence_pairs,
            random_off_energies,
        )

        print(f"Plot saved to: {random_output_plot}")
        print("Worst off-target interaction:")
        print(
            f"{random_strongest_offtarget['interaction_type']} | "
            f"energy = {random_strongest_offtarget['energy']:.3f} kcal/mol"
        )
        print(
            f"Pair A index {random_strongest_offtarget['pair_a_index']}: "
            f"{random_strongest_offtarget['pair_a'][0]} / "
            f"{random_strongest_offtarget['pair_a'][1]}"
        )
        print(
            f"Pair B index {random_strongest_offtarget['pair_b_index']}: "
            f"{random_strongest_offtarget['pair_b'][0]} / "
            f"{random_strongest_offtarget['pair_b'][1]}"
        )
        print(
            f"Interacting strands: "
            f"{random_strongest_offtarget['strand_a_type']}="
            f"{random_strongest_offtarget['strand_a']} vs "
            f"{random_strongest_offtarget['strand_b_type']}="
            f"{random_strongest_offtarget['strand_b']}"
        )
        print(random_stats)
        write_sequence_pairs_sheet(
            workbook,
            f"k={k_value} (random equivalent)",
            random_sequence_pairs,
        )

    workbook.save(workbook_path)
    print(f"Sequence workbook saved to: {workbook_path}")
