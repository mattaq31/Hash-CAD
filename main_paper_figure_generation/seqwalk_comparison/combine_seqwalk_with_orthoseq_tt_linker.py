"""Run the k=5 SeqWalk/legacy-search analysis with a 5' TT flank."""

from pathlib import Path
import random
import sys

import numpy as np
import openpyxl

PACKAGE_ROOT = Path(__file__).resolve().parents[2] / "crisscross_kit"
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))

from orthoseq_generator import helper_functions as hf
from orthoseq_generator import sequence_computations as sc
from orthoseq_generator import vertex_cover_algorithms as vca


DEFAULT_RANDOM_SEED = 41


def strongest_offtarget_interaction(sequence_pairs, off_energies):
    matrix_specs = [
        ("handle_handle", off_energies["handle_handle_energies"], "handle", "handle"),
        (
            "antihandle_antihandle",
            off_energies["antihandle_antihandle_energies"],
            "antihandle",
            "antihandle",
        ),
        (
            "handle_antihandle",
            off_energies["antihandle_handle_energies"],
            "handle",
            "antihandle",
        ),
    ]

    best = None
    for interaction_type, matrix, strand_type_a, strand_type_b in matrix_specs:
        valid_positions = np.argwhere(matrix != 0)
        for i, j in valid_positions:
            energy = matrix[i, j]
            if best is None or energy < best["energy"]:
                seq_a, rc_a = sequence_pairs[i]
                seq_b, rc_b = sequence_pairs[j]
                strand_a = seq_a if strand_type_a == "handle" else rc_a
                strand_b = seq_b if strand_type_b == "handle" else rc_b
                best = {
                    "interaction_type": interaction_type,
                    "energy": float(energy),
                    "pair_a_index": int(i),
                    "pair_b_index": int(j),
                    "pair_a": (seq_a, rc_a),
                    "pair_b": (seq_b, rc_b),
                    "strand_a_type": strand_type_a,
                    "strand_b_type": strand_type_b,
                    "strand_a": strand_a,
                    "strand_b": strand_b,
                }
    return best


def report_strongest_offtarget(sequence_pairs, off_energies):
    strongest_offtarget = strongest_offtarget_interaction(sequence_pairs, off_energies)
    if strongest_offtarget is None:
        print("Worst off-target interaction: none found")
        return

    print("Worst off-target interaction:")
    print(
        f"{strongest_offtarget['interaction_type']} | "
        f"energy = {strongest_offtarget['energy']:.3f} kcal/mol"
    )
    print(
        f"Pair A index {strongest_offtarget['pair_a_index']}: "
        f"{strongest_offtarget['pair_a'][0]} / {strongest_offtarget['pair_a'][1]}"
    )
    print(
        f"Pair B index {strongest_offtarget['pair_b_index']}: "
        f"{strongest_offtarget['pair_b'][0]} / {strongest_offtarget['pair_b'][1]}"
    )
    print(
        f"Interacting strands: "
        f"{strongest_offtarget['strand_a_type']}={strongest_offtarget['strand_a']} "
        f"vs {strongest_offtarget['strand_b_type']}={strongest_offtarget['strand_b']}"
    )


def write_sequence_pairs_sheet(workbook, sheet_name, sequence_pairs, on_energies=None):
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet["A1"] = sheet_name
    worksheet["A2"] = "ID"
    worksheet["B2"] = "Handle"
    worksheet["C2"] = "Antihandle"
    worksheet["D2"] = "On-target energy (kcal/mol)"

    if on_energies is None:
        on_energies = [None] * len(sequence_pairs)

    for idx, ((seq, rc_seq), on_energy) in enumerate(
        zip(sequence_pairs, on_energies),
        start=1,
    ):
        row = idx + 2
        worksheet.cell(row=row, column=1, value=idx)
        worksheet.cell(row=row, column=2, value=seq)
        worksheet.cell(row=row, column=3, value=rc_seq)
        if on_energy is not None:
            worksheet.cell(row=row, column=4, value=float(on_energy))


def post_select_by_ontarget(sequence_pairs, on_energies, max_ontarget_energy):
    selected_pairs = []
    selected_energies = []
    selected_original_indices = []

    for original_index, (pair, on_energy) in enumerate(zip(sequence_pairs, on_energies)):
        if on_energy < max_ontarget_energy:
            selected_pairs.append(pair)
            selected_energies.append(float(on_energy))
            selected_original_indices.append(original_index)

    return selected_pairs, np.array(selected_energies), selected_original_indices


def legacy_sequence_search(sequence_pairs, source_indices, off_energies, offtarget_limit):
    random.seed(DEFAULT_RANDOM_SEED + int(round((offtarget_limit + 100.0) * 1000)))

    id_to_seq = dict(zip(source_indices, sequence_pairs))
    edges = vca.build_edges(off_energies, source_indices, offtarget_limit)
    vertices = set(source_indices)

    print(
        f"Built legacy search graph with {len(vertices)} vertices and "
        f"{len(edges)} edges at off-target cutoff {offtarget_limit:.3f} kcal/mol."
    )

    vertex_cover, _trajectories = vca.iterative_vertex_cover_refinement(
        vertices,
        edges,
        avoid_V=None,
        num_vertices_to_remove=20,
        max_iterations=200,
        limit=70,
        show_progress=False,
    )

    independent_ids = sorted(vertices - vertex_cover)
    independent_sequences = [id_to_seq[i] for i in independent_ids]
    print(
        f"Legacy sequence search kept {len(independent_sequences)} of "
        f"{len(sequence_pairs)} post-selected sequence pairs."
    )
    return independent_sequences, independent_ids, edges


def iterative_legacy_sequence_search(
    sequence_pairs,
    source_indices,
    off_energies,
    initial_offtarget_limit,
    initial_delta,
    min_delta,
    target_sequence_count,
):
    cutoff = round(float(initial_offtarget_limit), 10)
    delta = float(initial_delta)
    best_result = None
    sweep_results = []
    max_search_rounds = 100
    search_round = 0

    while delta >= min_delta and search_round < max_search_rounds:
        search_round += 1
        print(f"\nRunning legacy sequence search at cutoff {cutoff:.3f} kcal/mol")
        legacy_pairs, legacy_source_indices, legacy_edges = legacy_sequence_search(
            sequence_pairs,
            source_indices,
            off_energies,
            cutoff,
        )
        sequence_count = len(legacy_pairs)
        sweep_results.append(
            {
                "offtarget_limit": cutoff,
                "sequence_count": sequence_count,
                "edge_count": len(legacy_edges),
                "delta": delta,
                "search_round": search_round,
                "distance_to_target": abs(sequence_count - target_sequence_count),
            }
        )

        candidate_result = (
            legacy_pairs,
            legacy_source_indices,
            cutoff,
            sweep_results.copy(),
        )
        if best_result is None:
            best_result = candidate_result
        else:
            best_count = len(best_result[0])
            current_distance = abs(sequence_count - target_sequence_count)
            best_distance = abs(best_count - target_sequence_count)
            if current_distance < best_distance:
                best_result = candidate_result
            elif current_distance == best_distance and sequence_count >= target_sequence_count:
                best_result = candidate_result

        if sequence_count == target_sequence_count:
            print(
                f"Found exact target of {target_sequence_count} sequences "
                f"at cutoff {cutoff:.3f}."
            )
            best_result = candidate_result
            break

        if sequence_count < target_sequence_count:
            print(
                f"{sequence_count} sequences is below target {target_sequence_count}; "
                f"moving cutoff down by {delta / 2:.3f}."
            )
            cutoff = round(cutoff - (delta / 2), 10)
            delta /= 2
            continue

        print(
            f"{sequence_count} sequences is above target {target_sequence_count}; "
            f"moving cutoff up by {delta:.3f}."
        )
        cutoff = round(cutoff + delta, 10)

    if best_result is None:
        print(
            "No legacy search round produced a result; no cutoff will be plotted."
        )
        return [], [], None, sweep_results

    legacy_pairs, legacy_source_indices, chosen_cutoff, _ = best_result
    print(
        f"Using cutoff {chosen_cutoff:.3f} kcal/mol with "
        f"{len(legacy_pairs)} legacy-selected sequences "
        f"(target {target_sequence_count})."
    )
    return legacy_pairs, legacy_source_indices, chosen_cutoff, sweep_results


def enforce_exact_sequence_count(
    sequence_pairs,
    source_indices,
    on_energy_by_source_index,
    target_sequence_count,
):
    sequence_count = len(sequence_pairs)
    if sequence_count == target_sequence_count:
        return sequence_pairs, source_indices

    if sequence_count < target_sequence_count:
        raise ValueError(
            f"Legacy search returned only {sequence_count} sequences; cannot make "
            f"an exact {target_sequence_count}-sequence set without adding "
            "conflicting sequences."
        )

    ranked = sorted(
        zip(sequence_pairs, source_indices),
        key=lambda item: (on_energy_by_source_index[item[1]], item[1]),
    )
    trimmed = ranked[:target_sequence_count]
    trimmed_pairs = [pair for pair, _source_index in trimmed]
    trimmed_source_indices = [source_index for _pair, source_index in trimmed]
    print(
        f"Trimmed legacy independent set from {sequence_count} to "
        f"exactly {target_sequence_count} sequences by keeping the strongest "
        "on-target energies."
    )
    return trimmed_pairs, trimmed_source_indices


if __name__ == "__main__":
    SEQ_LENGTH = 7
    K_VALUE = 5
    FIVEP_FLANK = "TT"
    RANDOM_SEED = DEFAULT_RANDOM_SEED
    MAX_ON_TARGET_ENERGY = -9.239
    OFFTARGET_LIMIT = MAX_ON_TARGET_ENERGY
    OFFTARGET_LIMIT_INITIAL_DELTA = 0.3
    OFFTARGET_LIMIT_MIN_DELTA = 0.005
    TARGET_LEGACY_SEQUENCE_COUNT = 64
    XLIM = (-13.5, 0.0)
    random.seed(RANDOM_SEED)

    output_dir = Path(
        "/Users/floriankatzmeier/Dropbox/CrissCross/Papers/hash_cad/"
        "Figures/final_figures/revision/figure_SB"
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    legacy_output_plot = (
        output_dir / "seqwalk_7mer_tt_on_off_target_hist_k5_legacy_selected.svg"
    )
    workbook_path = output_dir / "seqwalk_k5_tt_post_selected_sequence_pairs.xlsx"

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    hf.ENERGY_TYPE = "totalu"
    hf.set_nupack_params(material="dna", celsius=37, sodium=0.05, magnesium=0.025)

    sequence_pairs_pool = sc.create_seqwalk_sequence_pairs_pool(
        length=SEQ_LENGTH,
        k=K_VALUE,
        seed=RANDOM_SEED,
        fivep_ext=FIVEP_FLANK,
        threep_ext="",
        alphabet="ACGT",
        avoid_reverse_complements=True,
        prevented_patterns=[],
        verbose=True,
    )

    print(
        f"Generated {len(sequence_pairs_pool)} SeqWalk sequence pairs "
        f"for {SEQ_LENGTH}-mers at SSM k={K_VALUE} with 5' {FIVEP_FLANK} flanks."
    )

    sequence_pairs = [pair for _, pair in sequence_pairs_pool]
    sequence_count = len(sequence_pairs)

    on_energies, _, _ = sc.compute_ontarget_energies(sequence_pairs)
    write_sequence_pairs_sheet(workbook, f"k={K_VALUE} TT", sequence_pairs, on_energies)

    selected_pairs, selected_on_energies, selected_original_indices = (
        post_select_by_ontarget(
            sequence_pairs,
            on_energies,
            MAX_ON_TARGET_ENERGY,
        )
    )
    selected_count = len(selected_pairs)
    print(
        f"Post-selected {selected_count} of {sequence_count} sequence pairs "
        f"with on-target energy < {MAX_ON_TARGET_ENERGY:.3f} kcal/mol."
    )

    if selected_count == 0:
        print("No post-selected sequence pairs found; skipping legacy search and plot.")
        legacy_pairs = []
        legacy_on_energies = np.array([])
        legacy_source_indices = []
        chosen_offtarget_limit = None
        cutoff_sweep_results = []
    else:
        selected_off_energies = sc.compute_offtarget_energies(selected_pairs)
        (
            legacy_pairs,
            legacy_source_indices,
            chosen_offtarget_limit,
            cutoff_sweep_results,
        ) = iterative_legacy_sequence_search(
            selected_pairs,
            selected_original_indices,
            selected_off_energies,
            OFFTARGET_LIMIT,
            OFFTARGET_LIMIT_INITIAL_DELTA,
            OFFTARGET_LIMIT_MIN_DELTA,
            TARGET_LEGACY_SEQUENCE_COUNT,
        )
        selected_energy_by_source_index = dict(
            zip(selected_original_indices, selected_on_energies)
        )
        if len(legacy_pairs) > 0:
            legacy_pairs, legacy_source_indices = enforce_exact_sequence_count(
                legacy_pairs,
                legacy_source_indices,
                selected_energy_by_source_index,
                TARGET_LEGACY_SEQUENCE_COUNT,
            )
        legacy_on_energies = np.array(
            [selected_energy_by_source_index[i] for i in legacy_source_indices]
        )

        if len(legacy_pairs) == 0:
            print("Legacy sequence search selected no pairs; skipping final plot.")
        else:
            legacy_off_energies = sc.compute_offtarget_energies(legacy_pairs)
            legacy_stats = sc.plot_on_off_target_histograms(
                legacy_on_energies,
                legacy_off_energies,
                output_path=str(legacy_output_plot),
                show_plot=True,
                title="64 Post-Selected SeqWalk Pairs, 5' TT Flank",
                xlim=XLIM,
            )
            print(f"Plot saved to: {legacy_output_plot}")
            report_strongest_offtarget(legacy_pairs, legacy_off_energies)
            print(legacy_stats)

    write_sequence_pairs_sheet(
        workbook,
        f"k={K_VALUE} TT post-selected",
        selected_pairs,
        selected_on_energies,
    )

    index_sheet = workbook.create_sheet(title="post-selected source IDs")
    index_sheet["A1"] = "Selected row IDs from the full k=5 TT SeqWalk sheet"
    index_sheet["A2"] = "Post-selected ID"
    index_sheet["B2"] = "Full-sheet ID"
    for idx, original_index in enumerate(selected_original_indices, start=1):
        row = idx + 2
        index_sheet.cell(row=row, column=1, value=idx)
        index_sheet.cell(row=row, column=2, value=original_index + 1)

    write_sequence_pairs_sheet(
        workbook,
        f"k={K_VALUE} TT legacy-selected",
        legacy_pairs,
        legacy_on_energies,
    )

    legacy_index_sheet = workbook.create_sheet(title="legacy-selected source IDs")
    legacy_index_sheet["A1"] = "Selected row IDs from the full k=5 TT SeqWalk sheet"
    legacy_index_sheet["A2"] = "Legacy-selected ID"
    legacy_index_sheet["B2"] = "Full-sheet ID"
    for idx, original_index in enumerate(legacy_source_indices, start=1):
        row = idx + 2
        legacy_index_sheet.cell(row=row, column=1, value=idx)
        legacy_index_sheet.cell(row=row, column=2, value=original_index + 1)

    sweep_sheet = workbook.create_sheet(title="legacy cutoff sweep")
    sweep_sheet["A1"] = "Off-target cutoff (kcal/mol)"
    sweep_sheet["B1"] = "Legacy-selected sequence count"
    sweep_sheet["C1"] = "Graph edge count"
    sweep_sheet["D1"] = "Delta"
    sweep_sheet["E1"] = "Search round"
    sweep_sheet["F1"] = "Distance to target"
    sweep_sheet["G1"] = "Chosen"
    for row_idx, sweep_result in enumerate(cutoff_sweep_results, start=2):
        cutoff = sweep_result["offtarget_limit"]
        sweep_sheet.cell(row=row_idx, column=1, value=float(cutoff))
        sweep_sheet.cell(row=row_idx, column=2, value=sweep_result["sequence_count"])
        sweep_sheet.cell(row=row_idx, column=3, value=sweep_result["edge_count"])
        sweep_sheet.cell(row=row_idx, column=4, value=sweep_result["delta"])
        sweep_sheet.cell(row=row_idx, column=5, value=sweep_result["search_round"])
        sweep_sheet.cell(
            row=row_idx,
            column=6,
            value=sweep_result["distance_to_target"],
        )
        sweep_sheet.cell(
            row=row_idx,
            column=7,
            value=bool(
                chosen_offtarget_limit is not None
                and np.isclose(cutoff, chosen_offtarget_limit)
            ),
        )

    workbook.save(workbook_path)
    print(f"Sequence workbook saved to: {workbook_path}")
