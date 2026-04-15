"""
Load assembly handle pairs from Excel and plot on-target vs off-target energies.
"""

from pathlib import Path
import sys

import numpy as np
import openpyxl

PACKAGE_ROOT = Path(__file__).resolve().parents[2] / "crisscross_kit"
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))

from orthoseq_generator import helper_functions as hf
from orthoseq_generator import sequence_computations as sc


def strongest_offtarget_interaction(sequence_pairs, off_energies):
    matrix_specs = [
        ("handle_handle", off_energies["handle_handle_energies"], "handle", "handle"),
        (
            "antihandle_antihandle",
            off_energies["antihandle_antihandle_energies"],
            "antihandle",
            "antihandle",
        ),
        (
            "handle_antihandle",
            off_energies["antihandle_handle_energies"],
            "handle",
            "antihandle",
        ),
    ]

    best = None
    for interaction_type, matrix, strand_type_a, strand_type_b in matrix_specs:
        valid_positions = np.argwhere(matrix != 0)
        for i, j in valid_positions:
            energy = matrix[i, j]
            if best is None or energy < best["energy"]:
                seq_a, rc_a = sequence_pairs[i]
                seq_b, rc_b = sequence_pairs[j]
                strand_a = seq_a if strand_type_a == "handle" else rc_a
                strand_b = seq_b if strand_type_b == "handle" else rc_b
                best = {
                    "interaction_type": interaction_type,
                    "energy": float(energy),
                    "pair_a_index": int(i),
                    "pair_b_index": int(j),
                    "pair_a": (seq_a, rc_a),
                    "pair_b": (seq_b, rc_b),
                    "strand_a_type": strand_type_a,
                    "strand_b_type": strand_type_b,
                    "strand_a": strand_a,
                    "strand_b": strand_b,
                }
    return best


def load_sequence_pairs_from_excel(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[workbook.sheetnames[0]]

    sequence_pairs = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        if len(row) < 3 or row[1] is None or row[2] is None:
            continue
        sequence_pairs.append((str(row[1]).strip(), str(row[2]).strip()))
    return sequence_pairs


def strip_terminal_tt(sequence):
    sequence = sequence.strip().upper()
    if not sequence.startswith("TT"):
        raise ValueError(f"Expected sequence to start with 5' TT extension, got: {sequence}")
    return sequence[2:]


def remove_tt_extensions(sequence_pairs):
    trimmed_pairs = [
        (strip_terminal_tt(seq), strip_terminal_tt(rc_seq))
        for seq, rc_seq in sequence_pairs
    ]
    return trimmed_pairs


def report_strongest_offtarget(sequence_pairs, off_energies):
    strongest_offtarget = strongest_offtarget_interaction(sequence_pairs, off_energies)
    print("Worst off-target interaction:")
    print(
        f"{strongest_offtarget['interaction_type']} | "
        f"energy = {strongest_offtarget['energy']:.3f} kcal/mol"
    )
    print(
        f"Pair A index {strongest_offtarget['pair_a_index']}: "
        f"{strongest_offtarget['pair_a'][0]} / {strongest_offtarget['pair_a'][1]}"
    )
    print(
        f"Pair B index {strongest_offtarget['pair_b_index']}: "
        f"{strongest_offtarget['pair_b'][0]} / {strongest_offtarget['pair_b'][1]}"
    )
    print(
        f"Interacting strands: "
        f"{strongest_offtarget['strand_a_type']}={strongest_offtarget['strand_a']} "
        f"vs {strongest_offtarget['strand_b_type']}={strongest_offtarget['strand_b']}"
    )


if __name__ == "__main__":
    base_dir = Path(__file__).resolve().parent
    workbook_path = base_dir / "assembly_handle_pairs.xlsx"
    output_dir = Path(
        "/Users/floriankatzmeier/Dropbox/CrissCross/Papers/hash_cad/"
        "Figures/final_figures/revision/figure_SB"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    output_plot = output_dir / "assembly_handle_pairs_on_off_target_hist.svg"
    trimmed_output_plot = output_dir / "assembly_handle_pairs_no_tt_on_off_target_hist.svg"
    XLIM = (-13.5, 0.0)

    sequence_pairs = load_sequence_pairs_from_excel(workbook_path)
    trimmed_sequence_pairs = remove_tt_extensions(sequence_pairs)
    sequence_count = len(sequence_pairs)
    print(f"Loaded {len(sequence_pairs)} sequence pairs from {workbook_path.name}")
    changed_pairs = sum(
        1 for original, trimmed in zip(sequence_pairs, trimmed_sequence_pairs) if original != trimmed
    )
    print(f"Removed 5' TT extensions from {changed_pairs} sequence pairs.")
    hf.ENERGY_TYPE = "totalu"
    hf.USE_LIBRARY = False
    hf.set_nupack_params(material="dna", celsius=37, sodium=0.05, magnesium=0.025)

    on_energies, _, _ = sc.compute_ontarget_energies(sequence_pairs)
    off_energies = sc.compute_offtarget_energies(sequence_pairs)

    stats = sc.plot_on_off_target_histograms(
        on_energies,
        off_energies,
        output_path=str(output_plot),
        show_plot=True,
        title=f"{sequence_count} Assembly Handle Pairs",
        xlim=XLIM,
    )
    print(f"Plot saved to: {output_plot}")
    report_strongest_offtarget(sequence_pairs, off_energies)
    print(stats)

    trimmed_on_energies, _, _ = sc.compute_ontarget_energies(trimmed_sequence_pairs)
    trimmed_off_energies = sc.compute_offtarget_energies(trimmed_sequence_pairs)

    trimmed_stats = sc.plot_on_off_target_histograms(
        trimmed_on_energies,
        trimmed_off_energies,
        output_path=str(trimmed_output_plot),
        show_plot=True,
        title=f"{sequence_count} Assembly Handle Pairs Without TT Extensions",
        xlim=XLIM,
    )
    print(f"Plot saved to: {trimmed_output_plot}")
    report_strongest_offtarget(trimmed_sequence_pairs, trimmed_off_energies)
    print(trimmed_stats)
