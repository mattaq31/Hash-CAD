import openpyxl
import re
from collections import defaultdict
import os

def extract_regex_pattern(sheet):
    # Extract the regex pattern from the upper left-hand corner (cell A1)
    return sheet.cell(row=1, column=1).value

def parse_names(names, pattern):
    # Compile the regex pattern
    regex = re.compile(pattern)

    # Dictionary to store the first occurrence of each "id"
    id_to_word = {}

    for name in names:
        if(name):
            match = regex.match(name)
            if match:
                # Extract components from the match
                components = match.groups()
                # Assuming the "id" and the associated word are at specific positions
                id_num = components[1]  # Extract the "id" number
                word = components[0]  # Extract the associated word

                # Store the first occurrence of the "id"
                if id_num not in id_to_word:
                    id_to_word[id_num] = word

    return id_to_word


def import_plate_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)

    # Create a mapping of normalized sheet names to actual sheet objects
    sheet_map = {sheet.title.lower(): sheet for sheet in workbook}

    # Access sheets using normalized names
    name_sheet = sheet_map['names']
    sequence_sheet = sheet_map['sequences']
    description_sheet = sheet_map['descriptions']

    # Extract regex pattern from the name sheet
    regex_pattern = extract_regex_pattern(name_sheet)

    # Initialize the result list
    result = []

    # Collect all names
    all_names = []

    # Read the 16x24 grid with labels
    for row in range(2, 18):  # Rows 2 to 17 (1-based indexing)
        for col in range(2, 26):  # Columns B to Y (1-based indexing)
            well = f"{name_sheet.cell(row=row, column=1).value}{name_sheet.cell(row=1, column=col).value}"
            name = name_sheet.cell(row=row, column=col).value
            all_names.append(name)
            sequence = sequence_sheet.cell(row=row, column=col).value
            description = description_sheet.cell(row=row, column=col).value

            element = {
                'well': well,
                'name': name,
                'sequence': sequence,
                'description': description
            }
            result.append(element)

    # Parse the names using the extracted regex pattern
    cargo_key = parse_names(all_names, regex_pattern)

    return result, cargo_key, regex_pattern


class GenericCargoPlate:
    def __init__(self, cargo_key, plate_folder, plate_names, plates_data):
        self.cargo_key = cargo_key
        self.plate_folder = plate_folder
        self.plate_names = plate_names
        self.plates = plates_data
        self.sequences = defaultdict(bool)
        self.wells = defaultdict(bool)

    def add_sequence(self, position, side, name, sequence):
        self.sequences[(position, side, name)] = sequence

    def add_well(self, position, side, name, well):
        self.wells[(position, side, name)] = well


def createGenericPlate(plate_name, plate_folder):
    file_path = os.path.join(plate_folder, plate_name)

    plates_data, cargo_key, regex_pattern = import_plate_from_excel(file_path)

    # Create an instance of GenericCargoPlate
    plate = GenericCargoPlate(cargo_key, plate_folder, plate_name, plates_data)

    for item in plates_data:
        descriptor = item['name']
        sequence = item['sequence']
        well = item['well']

        position = 0
        side = 0
        name = ''

        # Compile the regex pattern
        regex = re.compile(regex_pattern)
        if(descriptor):
            match = regex.match(descriptor)
            if match:
                # Extract components from the match
                components = match.groups()
                position = components[3]  # Extract the "position" number
                side = components[2]  # Extract the "side" number
                name = components[0] # Extract the name

        plate.add_sequence(position, side, name, sequence)
        plate.add_well(position, side, name, well)

    return plate



#############################
#     Example usage         #
#############################
file_path = '/GUI/used-cargo-plates/sw_src007_nelson_quimby_bart_edna.xlsx'

#grid_data, cargo_key, regex = import_plate_from_excel(file_path)

# Print the result
#print("Grid Data:")
#for item in grid_data:
#    print(item)
#
#print("\nID to Word Map:")
#for id_num, word in cargo_key.items():
#    print(f"ID: {id_num}, Word: {word}")


plate_folder= 'C:\\Users\\cmbec\\OneDrive\\Cloud_Documents\\Shih_Lab_2024\\Crisscross-Design\\GUI\\used-cargo-plates'
plate_name = 'sw_src007_nelson_quimby_bart_edna.xlsx'

plate = createGenericPlate(plate_name, plate_folder)

print("Finished")