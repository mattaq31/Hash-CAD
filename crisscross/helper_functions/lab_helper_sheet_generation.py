from collections import defaultdict, OrderedDict
import numpy as np
from colorama import Fore
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import string
import math
from crisscross.plate_mapping.plate_concentrations import concentration_library


uppercase_alphabet = string.ascii_uppercase
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


def apply_box_border(ws, top_left, top_right, bottom_left, bottom_right, style='thick'):
    """
    Applies a thick border to an excel sheet surrounding the specified cells.
    :param ws: Excel worsheet object.
    :param top_left: Top left cell.
    :param top_right: Top right cell.
    :param bottom_left: Bottom left cell.
    :param bottom_right: Bottom right cell.
    :param style: Border style to be applied.
    :return: N/A, applied in-place.
    """

    selected_border = Side(border_style=style)

    # Top row
    for cell in ws[f"{top_left}:{top_right}"][0]:
        cell.border = Border(top=selected_border)
    # Bottom row
    for cell in ws[f"{bottom_left}:{bottom_right}"][0]:
        cell.border = Border(bottom=selected_border)
    # Left column
    for cell in ws[f"{top_left}:{bottom_left}"]:
        cell[0].border = Border(left=selected_border)
    # Right column
    for cell in ws[f"{top_right}:{bottom_right}"]:
        cell[0].border = Border(right=selected_border)

    # Top-left corner
    ws[top_left].border = Border(top=selected_border, left=selected_border)
    # Top-right corner
    ws[top_right].border = Border(top=selected_border, right=selected_border)
    # Bottom-left corner
    ws[bottom_left].border = Border(bottom=selected_border, left=selected_border)
    # Bottom-right corner
    ws[bottom_right].border = Border(bottom=selected_border, right=selected_border)

def adjust_column_width(ws):
    """
    Adjusts the column width of an excel sheet based on the maximum length of the content in each column.
    :param ws: Excel sheet object.
    :return: N/A, adjusted in-place.
    """
    # Adjust the column width based on the maximum length of the content in each column
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        # Find the maximum length of the content in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set the column width to the max_length
        adjusted_width = max_length
        ws.column_dimensions[column_letter].width = adjusted_width


def prepare_master_mix_sheet(slat_dict, echo_sheet=None, reference_handle_volume=150, reference_handle_concentration=500,
                             slat_mixture_volume=50, unique_transfer_volume_plates=None, workbook=None):
    """
    Prepares a 'master mix' sheet to be used for combining slat mixtures with scaffold
    and core staples into the final slat mixture.
    :param slat_dict: Dictionary of slats with slat names as keys and slat objects as values.
    :param echo_sheet: Exact list of commands sent to the Echo robot for this group of slats.
    :param reference_handle_volume: Reference staple volume for each handle in a pool in nL (this refers to the control handles plate).
    :param reference_handle_concentration: Reference staple concentration used for the core staples in uM (this refers to the control handles plate).
    All concentration values will be referenced to this value.
    :param slat_mixture_volume: Reaction volume (in uL) for a single slat annealing mixture.  Can be set to 'max' to use up all available handle mix.
    :param unique_transfer_volume_plates: Plates that have special non-standard volumes.
    This will be ignored if the echo sheet is provided with the exact details.
    :param workbook: The workbook to which the new excel sheet should be added.
    :return: Workbook with new sheet included.
    """

    slat_count = len(slat_dict)
    slat_concentration_distribution = []

    # calculate the minimum handle concentration for each slat (from the echo pools)

    # preference is to use echo sheet for calculation if available (as this has exceptions applied which won't be apparent in the slat dictionary)
    if echo_sheet is not None:
        for slat_name in slat_dict.keys():
            echo_indices = echo_sheet[echo_sheet['Component'].str.contains(fr"{slat_name}_h\d_", na=False)].index
            total_handle_mix_volume = sum(echo_sheet.loc[echo_indices]['Transfer Volume'].values)
            slat_concentration_nM = 1000 * (reference_handle_concentration * reference_handle_volume) / total_handle_mix_volume
            slat_concentration_distribution.append(slat_concentration_nM)
    else:
        for slat in slat_dict.values():
            total_handle_mix_volume = 0
            for handle in list(slat.H2_handles.values()) + list(slat.H5_handles.values()):
                if handle['plate'] in unique_transfer_volume_plates:
                    total_handle_mix_volume += unique_transfer_volume_plates[handle['plate']]
                else:
                    total_handle_mix_volume += reference_handle_volume * int(concentration_library['sw_src002'] / concentration_library[handle['plate']])
            slat_concentration_nM = 1000 * (reference_handle_concentration * reference_handle_volume) / total_handle_mix_volume
            slat_concentration_distribution.append(slat_concentration_nM)

    min_handle_mix_conc = min(slat_concentration_distribution)
    max_handle_mix_conc = max(slat_concentration_distribution)

    print(Fore.BLUE + f'Info: Lowest handle mixture concentration: {round(min_handle_mix_conc, 1)}nM, highest handle mixture concentration: {round(max_handle_mix_conc, 1)}nM.' + Fore.RESET)

    if (max_handle_mix_conc - min_handle_mix_conc) / min_handle_mix_conc > 0.1:
        print(Fore.MAGENTA + f'Warning: The handle mixtures generated have a wide concentration range. '
                             f'You could save on some staples by splitting the master mix preparation '
                             f'into two or more batches.' + Fore.RESET)

    if workbook is not None:
        wb = workbook
    else:
        wb = Workbook()

    ws = wb.create_sheet("Slat Folding & Master Mix")

    # Titles and formatting
    ws['A1'] = 'Single Slat Folding Quantities (can prepare directly or follow master mix details below)'
    ws['A1'].font = Font(bold=True)
    ws['A2'] = 'H2/H5 Handles from Echo [individual handle] (nM)'
    ws['A3'] = 'P8064 scaffold (nM)'
    ws['A4'] = 'Core staples pool (each, nM)'
    ws['A5'] = 'TEF (X)'
    ws['A6'] = 'MgCl2 (mM)'
    ws['A7'] = 'UPW (deionized water)'
    ws['A8'] = 'Total volume (µL)'
    ws['A9'] = 'Total amount (pmol)'

    # Standard concentration values
    ws['B1'] = 'Stock Concentration'
    ws['B2'] = round(min_handle_mix_conc, 1)
    ws['B3'] = 1062
    ws['B4'] = 3937
    ws['B5'] = 10
    ws['B6'] = 1000
    ws['B2'].fill = orange_fill
    ws['B3'].fill = red_fill

    # Standard target concentrations
    ws['C1'] = 'Final Concentration'
    ws['C2'] = 500
    ws['C3'] = 50
    ws['C4'] = 500
    ws['C5'] = 1
    ws['C6'] = 6

    # Calculations for single slat volumes
    ws['D1'] = 'Amount to Add (µL)'
    ws['D2'] = "=round(C2*D$8/B2,2)"
    ws['D3'] = "=round(C3*D$8/B3,2)"
    ws['D4'] = "=round(C4*D$8/B4,2)"
    ws['D5'] = "=round(C5*D$8/B5,2)"
    ws['D6'] = "=round(C6*D$8/B6,2)"
    ws['D7'] = '=D8-sum(D2:D6)'
    if slat_mixture_volume == "max":
        # selects the max reaction volume that reduces staple pool waste, assuming echo transfers only 75% of expected
        # 0.75 * total femtomoles of slats (based on min conc)/reaction final conc (500 nM), rounded down to the nearest multiple of 5
        ws['D8'] = math.floor(0.75 * min_handle_mix_conc * total_handle_mix_volume/ws['C2'].value/5/1000) * 5
        print(Fore.BLUE + f'Info: You selected "max", so the slat mixtures will use all the staples in your pooled handle mixtures.  The output slat mixture volume will be {ws["D8"].value}μl.' + Fore.RESET)

    else:
        ws['D8'] = slat_mixture_volume # simplest solution: user decides, defaults to 50 µL

    ws['D9'] = '=D8*C3/1000'

    ws['F2'].fill = red_fill
    ws['G2'] = 'Cells with this shading should be adjusted to match the actual values in your experiment.'
    ws['F3'].fill = orange_fill
    ws['G3'] = 'Cells with this shading contain an average or minimum value for a group of slats - if further precision is required for each slat, this needs to be changed.'

    apply_box_border(ws, 'A1', 'D1', 'A9', 'D9')

    # Calculations for master mix
    ws['A11'] = 'Master Mix Preparation (prepare once)'
    ws['A11'].font = Font(bold=True)
    ws['B11'] = 'Count or Volume (µL)'
    ws['A12'] = 'Number of slats (with a buffer of 3 extra slats)'

    ws['A13'] = 'P8064 scaffold'
    ws['A14'] = 'Core staples pool'
    ws['A15'] = 'TEF'
    ws['A16'] = 'MgCl2'
    ws['A17'] = 'UPW (deionized water)'
    ws['A18'] = 'Total volume'

    ws['B12'] = slat_count + 3
    ws['B13'] = '=D3*B12'
    ws['B14'] = '=D4*B12'
    ws['B15'] = '=D5*B12'
    ws['B16'] = '=D6*B12'
    ws['B17'] = '=D7*B12'
    ws['B18'] = '=SUM(B13:B17)'

    apply_box_border(ws, 'A11', 'B11', 'A18', 'B18')

    special_border = Side(border_style='mediumDashDotDot')
    thick_border = Side(border_style='thick')

    ws['A12'].border = Border(top=special_border, left=thick_border, bottom=special_border)
    ws['B12'].border = Border(top=special_border, right=thick_border, bottom=special_border)

    # Calculations for master mix + handle mix
    ws['A20'] = 'Final Slat Mixture (prepare once for each slat)'
    ws['A20'].font = Font(bold=True)
    ws['B20'] = 'Volume (µL)'

    ws['A21'] = 'Master Mix'
    ws['A22'] = 'Slat Handle Mixture'
    ws['A23'] = 'Total volume'

    ws['B21'] = '=SUM(B13:B17)/B12'
    ws['B22'] = '=D2'
    ws['B23'] = '=B21+B22'

    apply_box_border(ws, 'A20', 'B20', 'A23', 'B23')

    adjust_column_width(ws)

    return wb

def prepare_peg_purification_sheet(slat_dict, groups_per_layer=2, max_slat_concentration_uM=2,
                                   slat_mixture_volume=50,
                                   workbook=None, echo_sheet=None, special_slat_groups=None):
    """
    Prepares standard instructions for combining and purifying slat mixtures using PEG purification.  Also prepares lists of slat groups as a reference for when in the lab.
    :param slat_dict: Dictionary of slats with slat names as keys and slat objects as values.
    :param groups_per_layer: Number of PEG groups to use per crisscross layer.  You might want to adjust this if you have too many slats together in one group.
    :param max_slat_concentration_uM: Maximum concentration of slats in a combined PEG mixture (in UM) before a warning is triggered.
    :param slat_mixture_volume: Reaction volume (in uL) for a single slat annealing mixture.
    :param workbook: The workbook to which the new excel sheet should be added.
    :param echo_sheet: Exact list of commands sent to the Echo robot for this group of slats.
    :param special_slat_groups: IDs of slats that should be separated from the general slat groups and placed in their own group.
    :return: Workbook with new sheet included.
    """
    if workbook is not None:
        wb = workbook
    else:
        wb = Workbook()

    ws = wb.create_sheet("PEG Purification")

    layer_groups = defaultdict(list)
    full_data_groups = OrderedDict()

    for slat in slat_dict.values():
        if slat.ID not in layer_groups[slat.layer]:
            layer_groups[slat.layer].append(slat.ID)

    # Assigns slats to a group based on the number of groups per layer
    for layer, slats in layer_groups.items():
        if len(slats) < groups_per_layer:
            full_data_groups[f'L{layer}-ALL'] = {'IDs': slats}
        else:
            start_point = 0
            number_jump = len(slats) // groups_per_layer
            for i in range(1, groups_per_layer+1):
                full_data_groups[f'L{layer}-G{i}'] = {'IDs': slats[start_point:start_point+number_jump]}
                start_point += number_jump

    if special_slat_groups is not None:
        for special_group, slats in special_slat_groups.items():
            for slat in slats:
                for standard_group in full_data_groups:
                    if slat in full_data_groups[standard_group]['IDs']:
                        full_data_groups[standard_group]['IDs'].remove(slat)
            full_data_groups[special_group] = {'IDs': slats}

    # cleaning up empty groups
    keys_to_remove = []
    for position, group in enumerate(full_data_groups.keys()):
        if len(full_data_groups[group]['IDs']) == 0:  # this means that all slats in this group were moved to a special group and so this group is now empty
            keys_to_remove.append(group)
    for key in keys_to_remove:
        del full_data_groups[key]

    # Titles and formatting
    ws['A1'] = 'PEG Purification'
    ws['A1'].font = Font(bold=True)
    ws['A2'] = 'Step 0: Combine all slats into groups (slat IDs for each group beneath table)'
    ws['A3'] = 'Volume extracted from each well (µl)'
    ws['A4'] = '# of slats'
    ws['A5'] = 'Total volume expected (µL)'
    ws['A6'] = 'Scaffold Concentration per slat (nM)'
    ws['A7'] = 'Expected total origami amount (pmol, all)'
    ws['A8'] = 'Original Mg conc (mM)'
    ws['A9'] = 'Target Final Mg conc (mM)'
    ws['A10'] = 'Step 1: Add 1M Mg'
    ws['A11'] = 'Amount of 1M Mg to add (µl)'
    ws['A12'] = 'Step 2: Add 2X PEG'
    ws['A13'] = 'Amount of 2X PEG to add (µl)'
    ws['A14'] = 'Step 3: SPIN FOR 30 MINS AT 16KG, RT'
    ws['A15'] = 'Step 4: REMOVE SUPERNATANT AND ADD 150ul of RESUS1'
    ws['A16'] = 'Step 5: SPIN AGAIN FOR 30 MINS AT 16KG, RT'
    ws['A17'] = 'Step 6: REMOVE SUPERNATANT AND RESUSPEND IN RESUS2 AS BELOW'
    ws['A18'] = 'Desired final concentration (nM, per slat)'
    ws['A19'] = 'Resuspend with Resus2 to achieve target concentration for each slat (µl)'
    ws['A20'] = 'Expected total slat concentration (µM)'
    ws['A21'] = 'Step 7: SHAKE AT 33C FOR 1 HOUR AT 1000RPM, THEN NANODROP'
    ws['A22'] = 'Final Nanodrop (1x dilution - ng/µl dsDNA)'
    ws['A23'] = 'Average slat molecular weight (Da)'
    ws['A24'] = 'Total concentration from Nanodrop (µM)'
    ws['A25'] = 'Estimated concentration of each individual slat (nM)'
    ws['A26'] = 'Total amount of each slat (pmol)'
    ws['A27'] = 'Total origami (pmol)'
    ws['A28'] = 'PEG Yield (%)'

    # merge and center
    for cell in ['A2', 'A10', 'A12', 'A14', 'A15', 'A16', 'A17', 'A21']:
        ws.merge_cells(f'{cell}:{uppercase_alphabet[len(full_data_groups)]}{cell[1:]}')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].font = Font(bold=True)

    # fills in the equations and data for each group
    for position, group in enumerate(full_data_groups.keys()):
        column = uppercase_alphabet[position+1]
        # block 1 - slat counts and volumes
        full_data_groups[group][f'{column}1'] = group
        try:
            full_data_groups[group][f'{column}3'] = wb["Slat Folding & Master Mix"]["D8"].value
        except KeyError:
            print("No Slat Folding & Master Mix sheet detected. Defaulting to 50 µL reaction volume...")
            full_data_groups[group][f'{column}3'] = slat_mixture_volume
        full_data_groups[group][f'{column}4'] = len(full_data_groups[group]['IDs'])
        full_data_groups[group][f'{column}5'] = f"={column}3*{column}4"
        full_data_groups[group][f'{column}6'] = 50
        ws[f'{column}6'].fill = orange_fill

        # block 2 - magnesium addition
        full_data_groups[group][f'{column}7'] = f"={column}5*{column}6/1000"
        full_data_groups[group][f'{column}8'] = 6
        ws[f'{column}8'].fill = orange_fill
        full_data_groups[group][f'{column}9'] = 20
        full_data_groups[group][f'{column}11'] = f"=round(({column}9-{column}8)*{column}5/(1000-{column}9),2)"
        full_data_groups[group][f'{column}13'] = f"={column}11 + {column}5"

        # block 3 - resuspension calculations
        full_data_groups[group][f'{column}18'] = 100
        ws[f'{column}18'].fill = blue_fill
        full_data_groups[group][f'{column}19'] = f"=ROUND((({column}7/{column}4)/{column}18)*1000,1)"
        full_data_groups[group][f'{column}20'] = f"=({column}18 * {column}4)/1000"
        rule = CellIsRule(operator='greaterThan', formula=[f'{max_slat_concentration_uM}'], fill=red_fill)
        ws.conditional_formatting.add(f'{column}20', rule)
        ws[f'{column}22'].fill = green_fill

        # block 4 - MW and concentration calculations
        mw_total = 0
        for id in full_data_groups[group]['IDs']:
            mw_total += slat_dict[id].get_molecular_weight()
        full_data_groups[group][f'{column}23'] = mw_total / len(full_data_groups[group]['IDs'])

        full_data_groups[group][f'{column}24'] = f'=round(({column}22*1000)/{column}23,2)'
        full_data_groups[group][f'{column}25'] = f'=round({column}24/{column}4*1000,2)'
        full_data_groups[group][f'{column}26'] = f'=round({column}25*{column}19/1000,2)'
        full_data_groups[group][f'{column}27'] = f'={column}26*{column}4'
        full_data_groups[group][f'{column}28'] = f'={column}27/{column}7*100'

    # sidebar definitions
    sidebar_col_start = uppercase_alphabet[position+3]
    sidebar_col_2 = uppercase_alphabet[position + 4]
    ws[f'{sidebar_col_start}3'].fill = red_fill
    ws[f'{sidebar_col_2}3'] = 'If these cells are red, then your slat mixture is over the 2µM limit - there is a high chance the mixture will aggregate.'
    ws[f'{sidebar_col_start}5'].fill = blue_fill
    ws[f'{sidebar_col_2}5'] = 'Change the target concentration of these cells to reduce the concentration below the 2µM limit.'

    ws[f'{sidebar_col_start}4'].fill = orange_fill
    ws[f'{sidebar_col_2}4'] = 'If for any reason slat folding conditions are changed, make sure to update these cells.'

    ws[f'{sidebar_col_start}6'].fill = green_fill
    ws[f'{sidebar_col_2}6'] = 'Fill these cells with your nanodrop values.  It is suggested to dilute by 10 for nanodrop if you have more than 5 slats in one mixture.'

    # fills out the data for each marked cell
    for _, group in full_data_groups.items():
        for cell, value in group.items():
            if cell != 'IDs':
                ws[cell] = value

    # resus 1/2 handy values
    ws['A30'] = 'Resus1/2 Buffer Components'
    ws['A30'].font = Font(bold=True)
    ws['A31'] = '10X TEF'
    ws['A32'] = 'MgCl2 (mM)'
    ws['A33'] = 'UPW (deionized water)'
    ws['A34'] = 'Total Volume'
    ws['B30'] = 'Stock'
    ws['C30'] = 'Resus 1'
    ws['D30'] = 'Resus 2'

    ws['B31'] = 10
    ws['B32'] = 1000
    ws['C34'] = 2000
    ws['D34'] = 2000

    ws['C31'] = '=1*C34/B31'
    ws['D31'] = '=1*D34/B31'

    ws['C32'] = '=ROUND(C34*20/B32,2)'
    ws['D32'] = '=ROUND(D34*10/B32,2)'

    ws['C33'] = '=C34 - C32 - C31'
    ws['D33'] = '=D34 - D32 - D31'

    apply_box_border(ws, 'A30', 'D30', 'A34', 'D34')
    apply_box_border(ws, 'A1', f'{column}1', 'A28', f'{column}28')

    # slat group components and values
    ws['A36'] = 'Slat Group Components'
    ws.merge_cells(f'A36:D36')
    ws['A36'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A36'].font = Font(bold=True)

    ws['A37'] = 'Group Name'
    ws['A37'].font = Font(bold=True)
    ws['B37'] = 'Slat ID'
    ws['B37'].font = Font(bold=True)
    ws['C37'] = 'Slat Well'
    ws['C37'].font = Font(bold=True)
    ws['D37'] = 'Plate Name'
    ws['D37'].font = Font(bold=True)

    indexer = 38
    divider = Side(border_style='thick')
    for group in full_data_groups.keys():
        for id in full_data_groups[group]['IDs']:
            ws[f'A{indexer}'] = group
            ws[f'B{indexer}'] = id
            # if echo data available, can also point towards the exact plate wells
            if echo_sheet is not None:
                echo_index = echo_sheet[echo_sheet['Component'].str.contains(fr"{id}_h\d_", na=False)].index[0]
                ws[f'C{indexer}'] = echo_sheet.loc[echo_index]['Destination Well']
                ws[f'D{indexer}'] = echo_sheet.loc[echo_index]['Destination Plate Name']
            indexer += 1
        for cell in ws[f"A{indexer-1}:D{indexer-1}"][0]:
            cell.border = Border(bottom=divider)

    adjust_column_width(ws)

    return wb

def prepare_all_standard_sheets(slat_dict, save_filepath, reference_single_handle_volume=150,
                                reference_single_handle_concentration=500,
                                slat_mixture_volume=50,
                                peg_groups_per_layer=2,
                                echo_sheet=None,
                                max_slat_concentration_uM=2,
                                unique_transfer_volume_plates=None,
                                special_slat_groups=None):
    """
    Prepares a series of excel sheets to aid lab assembler while preparing and purifying slat mixtures.
    :param slat_dict: Dictionary of slats to be assembled (each item in the dict is a Slat Object containing all 64 handles in place)
    :param save_filepath: Output file path for the combined excel workbook
    :param reference_single_handle_volume: Reference staple volume for each handle in a pool in nL (this refers to the control handles plate).
    :param reference_single_handle_concentration: Reference staple concentration used for the core staples in uM (this refers to the control handles plate).
    All concentration values will be referenced to this value.
    :param slat_mixture_volume: Reaction volume (in uL) for a single slat annealing mixture. Can be set to 'max' to use up all available handle mix.
    :param peg_groups_per_layer: Number of PEG groups to use per crisscross layer.  You might want to adjust this if you have too many slats together in one group.
    :param echo_sheet: Exact echo commands to use as a reference for calculating slat concentrations.
    :param max_slat_concentration_uM: Maximum concentration of slats in a combined PEG mixture (in UM) before a warning is triggered.
    :param unique_transfer_volume_plates: Plates that have special non-standard volumes.  This will be ignored if the echo sheet is provided with the exact details.
    :param special_slat_groups: IDs of slats that should be separated from the general slat groups and placed in their own group.
    :return: N/A, file saved directly to disk.
    """


    wb = Workbook()
    wb.remove(wb["Sheet"])

    # prepares slat assembly mixture details
    prepare_master_mix_sheet(slat_dict, echo_sheet, reference_single_handle_volume, reference_single_handle_concentration,
                             slat_mixture_volume, unique_transfer_volume_plates, wb)

    # prepares slat purification details
    prepare_peg_purification_sheet(slat_dict, peg_groups_per_layer, max_slat_concentration_uM, slat_mixture_volume,
                                   wb, echo_sheet=echo_sheet, special_slat_groups=special_slat_groups)
    wb.save(save_filepath)
